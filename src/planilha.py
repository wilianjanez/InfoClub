"""
Leitura e atualização da planilha de temas (data/temas.xlsx).
"""
import random
from datetime import datetime
from openpyxl import load_workbook


def sortear_tema(caminho_planilha: str) -> dict:
    """
    Abre a planilha, filtra os temas com Publicado != True,
    sorteia um aleatoriamente e retorna seus dados.
    Lança RuntimeError se não houver temas disponíveis.
    """
    wb = load_workbook(caminho_planilha)
    sheet = wb.active

    disponiveis = []
    for row_idx in range(2, sheet.max_row + 1):
        publicado = sheet.cell(row=row_idx, column=4).value
        if publicado is True or str(publicado).strip().upper() == "TRUE":
            continue
        tema_id = sheet.cell(row=row_idx, column=1).value
        tema = sheet.cell(row=row_idx, column=2).value
        categoria = sheet.cell(row=row_idx, column=3).value
        if tema_id is None or tema is None:
            continue
        disponiveis.append(
            {"row": row_idx, "id": tema_id, "tema": tema, "categoria": categoria}
        )

    if not disponiveis:
        raise RuntimeError(
            "Nenhum tema disponível na planilha (todos já estão marcados como Publicado)."
        )

    return random.choice(disponiveis)


def marcar_como_publicado(caminho_planilha: str, row_idx: int) -> None:
    """
    Marca a linha indicada como Publicado=True e registra a data/hora.
    Salva o arquivo no mesmo caminho.
    """
    wb = load_workbook(caminho_planilha)
    sheet = wb.active

    sheet.cell(row=row_idx, column=4).value = True
    sheet.cell(row=row_idx, column=5).value = datetime.utcnow().strftime(
        "%Y-%m-%d %H:%M:%S UTC"
    )

    wb.save(caminho_planilha)
