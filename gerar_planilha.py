from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
sheet = wb.active
sheet.title = "Temas"

headers = ["ID", "Tema", "Categoria", "Publicado", "DataPublicacao"]
sheet.append(headers)

header_font = Font(bold=True, color="FFFFFF", name="Arial")
header_fill = PatternFill("solid", start_color="2F5597")
for col_idx, _ in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

temas_exemplo = [
    (1, "Como funciona a reciclagem de plástico", "Meio Ambiente", False, ""),
    (2, "Os 5 sentidos do corpo humano", "Ciência", False, ""),
    (3, "Como surgiu o café", "Curiosidades", False, ""),
    (4, "Diferença entre clima e tempo", "Meio Ambiente", False, ""),
    (5, "Como funciona a vacina", "Saúde", False, ""),
    (6, "A história da internet", "Tecnologia", False, ""),
    (7, "Por que o céu é azul", "Ciência", False, ""),
    (8, "Como economizar energia em casa", "Sustentabilidade", False, ""),
    (9, "Benefícios de beber água", "Saúde", False, ""),
    (10, "Como funciona um eclipse", "Astronomia", False, ""),
]

for row in temas_exemplo:
    sheet.append(list(row))

for row_idx in range(2, sheet.max_row + 1):
    sheet.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")

dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
sheet.add_data_validation(dv)
dv.add(f"D2:D1000")

widths = {"A": 6, "B": 45, "C": 20, "D": 12, "E": 18}
for col, width in widths.items():
    sheet.column_dimensions[col].width = width

for col_idx in range(1, 6):
    for row_idx in range(2, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=col_idx).font = Font(name="Arial")

sheet.freeze_panes = "A2"

wb.save("data/temas.xlsx")
print("Planilha criada em data/temas.xlsx")
