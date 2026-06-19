# -*- coding: utf-8 -*-
"""
Adiciona os novos temas (países, animais específicos, mecânicos populares) à
planilha data/temas.xlsx EXISTENTE, sem apagar o que já está lá.
Evita duplicar temas que já existam (por texto exato) e continua a sequência de ID.
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from gerar_temas_extra_v2 import gerar_novos_temas

CAMINHO_PLANILHA = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "data", "temas.xlsx"
)

wb = load_workbook(CAMINHO_PLANILHA)
sheet = wb.active

temas_existentes = set()
max_id = 0
for row_idx in range(2, sheet.max_row + 1):
    tema_id = sheet.cell(row=row_idx, column=1).value
    tema_texto = sheet.cell(row=row_idx, column=2).value
    if tema_id is not None:
        max_id = max(max_id, int(tema_id))
    if tema_texto:
        temas_existentes.add(tema_texto)

novos_temas = gerar_novos_temas()

adicionados = 0
ignorados_duplicados = 0
proximo_id = max_id + 1

for tema_texto, categoria in novos_temas:
    if tema_texto in temas_existentes:
        ignorados_duplicados += 1
        continue
    sheet.append([proximo_id, tema_texto, categoria, False, ""])
    row_idx = sheet.max_row
    sheet.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
    for col_idx in range(1, 6):
        sheet.cell(row=row_idx, column=col_idx).font = Font(name="Arial")
    temas_existentes.add(tema_texto)
    proximo_id += 1
    adicionados += 1

last_row = sheet.max_row

# Reaplica a validação de dados (TRUE/FALSE) cobrindo todas as linhas, incluindo as novas
for dv_existente in list(sheet.data_validations.dataValidation):
    sheet.data_validations.dataValidation.remove(dv_existente)

dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
sheet.add_data_validation(dv)
dv.add(f"D2:D{last_row}")

wb.save(CAMINHO_PLANILHA)

print(f"Temas adicionados: {adicionados}")
print(f"Temas ignorados (já existiam): {ignorados_duplicados}")
print(f"Total de temas na planilha agora: {last_row - 1}")
