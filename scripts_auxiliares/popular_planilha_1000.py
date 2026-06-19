import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from gerar_temas_lista import gerar_lista_completa

wb = Workbook()
sheet = wb.active
sheet.title = "Temas"

headers = ["ID", "Tema", "Categoria", "Publicado", "DataPublicacao"]
sheet.append(headers)

header_font = Font(bold=True, color="FFFFFF", name="Arial")
header_fill = PatternFill("solid", start_color="2F5597")
for col_idx, _ in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

temas = gerar_lista_completa(1000)

for idx, (tema, categoria) in enumerate(temas, start=1):
    sheet.append([idx, tema, categoria, False, ""])

last_row = sheet.max_row

for row_idx in range(2, last_row + 1):
    sheet.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
    for col_idx in range(1, 6):
        sheet.cell(row=row_idx, column=col_idx).font = Font(name="Arial")

dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
sheet.add_data_validation(dv)
dv.add(f"D2:D{last_row}")

widths = {"A": 6, "B": 55, "C": 25, "D": 12, "E": 18}
for col, width in widths.items():
    sheet.column_dimensions[col].width = width

sheet.freeze_panes = "A2"

caminho_saida = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "temas.xlsx")
wb.save(caminho_saida)
print(f"Planilha gerada com {last_row - 1} temas em data/temas.xlsx")
